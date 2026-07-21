"""Tests for the multi-tab input workbook template."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from models import DataParser
from ingestion import DATA_ENTRY_PROFILE_ID, FormatDetector, ingest_workbook
from workbook_template import (
    build_input_template,
    build_input_template_bytes,
    load_data_entry_sheets,
    load_project_metadata,
)


EXPECTED_SHEETS = {
    "Instructions",
    "Project",
    "Collars",
    "Lithology",
    "Water",
    "Environmental",
    "Screens",
    "Gradients",
    "Example",
    "Data Entry",
}


def test_build_input_template_writes_multi_tabs(tmp_path: Path) -> None:
    path = build_input_template(tmp_path / "template.xlsx")
    import pandas as pd

    workbook = pd.ExcelFile(path)
    assert EXPECTED_SHEETS.issubset(set(workbook.sheet_names))
    assert workbook.sheet_names[0] == "Instructions"
    assert workbook.sheet_names[1] == "Project"
    assert workbook.sheet_names[-1] == "Data Entry"


def test_build_input_template_bytes_matches_sheet_set() -> None:
    from io import BytesIO

    import pandas as pd

    payload = build_input_template_bytes()
    assert isinstance(payload, (bytes, bytearray))
    assert len(payload) > 1000
    workbook = pd.ExcelFile(BytesIO(payload))
    assert EXPECTED_SHEETS.issubset(set(workbook.sheet_names))


def test_format_detector_recognizes_data_entry_template(tmp_path: Path) -> None:
    path = build_input_template(tmp_path / "template.xlsx")
    detection = FormatDetector().detect(path)
    assert detection.profile_id == DATA_ENTRY_PROFILE_ID
    assert detection.confidence == 1.0


def test_ingest_workbook_data_entry_template(tmp_path: Path) -> None:
    path = build_input_template(tmp_path / "template.xlsx")
    result, report = ingest_workbook(path)
    assert report.profile_id == DATA_ENTRY_PROFILE_ID
    assert len(result.collars) == 3
    assert len(result.lithologies) >= 8
    assert len(result.environmental_readings) == 5
    assert "Data Entry" in report.optional_sheets_detected
    assert report.project_metadata["client_name"] == "C-GROUP ENERGY INC."
    assert report.project_metadata["prepared_by"] == "ECOVENTURE"


def test_data_entry_template_parses_in_dataparser(tmp_path: Path) -> None:
    path = build_input_template(tmp_path / "template.xlsx")
    result = DataParser().parse_file(path)
    assert len(result.collars) == 3
    assert len(result.lithologies) >= 8
    assert len(result.water_levels) == 3
    assert len(result.environmental_readings) == 5
    assert len(result.screen_intervals) == 2
    assert len(result.vertical_gradients) == 2
    chloride = [reading for reading in result.environmental_readings if reading.parameter == "Chloride"]
    assert len(chloride) == 4
    assert any(reading.value_label == "<5 mg/L" for reading in chloride)


def test_load_data_entry_project_metadata(tmp_path: Path) -> None:
    path = build_input_template(tmp_path / "template.xlsx")
    sheets = load_data_entry_sheets(path)
    assert sheets.project["client_name"] == "C-GROUP ENERGY INC."
    assert sheets.project["section_title"].startswith("A - A")
    project = load_project_metadata(path)
    assert project["client_name"] == "C-GROUP ENERGY INC."
    assert project["map_scale"] == "1:1000"


def test_example_sheet_is_not_required_for_parse(tmp_path: Path) -> None:
    """Example is reference-only; geology comes from named Collars/Lithology tabs."""
    path = build_input_template(tmp_path / "template.xlsx")
    import pandas as pd
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    workbook.remove(workbook["Example"])
    workbook.save(path)
    result = DataParser().parse_file(path)
    assert len(result.collars) == 3
    collars = pd.read_excel(path, sheet_name="Collars")
    assert "MW-01" in set(collars["hole_id"].astype(str))


def test_hybrid_native_plus_data_entry_keeps_environmental(tmp_path: Path) -> None:
    path = build_input_template(tmp_path / "hybrid.xlsx")
    # Template already has native Collars/Lithology; overlays come from those sheets
    # (Data Entry geology sections are empty stubs).
    hybrid = DataParser().parse_file(path)
    assert len(hybrid.collars) == 3
    assert len(hybrid.environmental_readings) == 5
    assert len(hybrid.water_levels) == 3
    # Ensure Data Entry still carries PROJECT for metadata path
    sheets = load_data_entry_sheets(path)
    assert sheets.project["client_name"] == "C-GROUP ENERGY INC."
    assert sheets.collars.empty
    assert sheets.lithology.empty


def test_ingest_native_multi_tab_keeps_overlays(tmp_path: Path) -> None:
    """Native detect path must not drop Water/Environmental when Collars exist."""
    from openpyxl import load_workbook

    path = build_input_template(tmp_path / "native_overlays.xlsx")
    workbook = load_workbook(path)
    # Remove Data Entry so FormatDetector chooses native_platform.
    workbook.remove(workbook["Data Entry"])
    workbook.save(path)
    detection = FormatDetector().detect(path)
    assert detection.profile_id == "native_platform"
    result, _report = ingest_workbook(path)
    assert len(result.collars) == 3
    assert len(result.water_levels) == 3
    assert len(result.environmental_readings) == 5
    assert len(result.screen_intervals) == 2


def test_load_project_metadata_merges_partial_project(tmp_path: Path) -> None:
    import pandas as pd

    path = build_input_template(tmp_path / "merge_meta.xlsx")
    # Wipe most Project values but keep one field — Data Entry should fill the rest.
    frame = pd.read_excel(path, sheet_name="Project")
    frame.loc[frame["field"] != "notes", "value"] = ""
    frame.loc[frame["field"] == "notes", "value"] = "Only notes on Project tab"
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        frame.to_excel(writer, sheet_name="Project", index=False)
    project = load_project_metadata(path)
    assert project["notes"] == "Only notes on Project tab"
    assert project["client_name"] == "C-GROUP ENERGY INC."
