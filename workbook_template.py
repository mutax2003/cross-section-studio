"""Multi-tab Cross Section Studio input workbook.

Sheets:
  Instructions — how to fill the workbook
  Project — client / figure metadata (also mirrored on Data Entry for ingest)
  Collars, Lithology, Water, Environmental, Screens, Gradients — primary data entry
  Example — filled sample project (reference only; not parsed)
  Data Entry — PROJECT block for ``DATA_ENTRY_PROFILE_ID`` + consulting seed
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

INSTRUCTIONS_SHEET = "Instructions"
PROJECT_SHEET = "Project"
DATA_ENTRY_SHEET = "Data Entry"
EXAMPLE_SHEET = "Example"

COLLARS_SHEET = "Collars"
LITHOLOGY_SHEET = "Lithology"
WATER_SHEET = "Water"
ENVIRONMENTAL_SHEET = "Environmental"
SCREENS_SHEET = "Screens"
GRADIENTS_SHEET = "Gradients"

DATA_SHEETS: tuple[str, ...] = (
    COLLARS_SHEET,
    LITHOLOGY_SHEET,
    WATER_SHEET,
    ENVIRONMENTAL_SHEET,
    SCREENS_SHEET,
    GRADIENTS_SHEET,
)

PROJECT_FIELDS: tuple[tuple[str, str], ...] = (
    ("client_name", "Client / prepared-for organization"),
    ("prepared_by", "Consultant or author firm"),
    ("project_number", "Project or file number"),
    ("section_title", "Cross-section title (e.g. A - A' WITH CHLORIDE)"),
    ("report_date", "Report date (DD/MM/YY)"),
    ("drawn_by", "Drawn-by initials"),
    ("data_source", "Data source citation"),
    ("map_scale", "Map scale (e.g. 1:1000)"),
    ("coordinate_reference", "CRS (e.g. EPSG:32611)"),
    ("transect_start", "Transect start label (e.g. A / NORTHWEST)"),
    ("transect_end", "Transect end label (e.g. A' / SOUTHEAST)"),
    ("vertical_exaggeration", "Suggested vertical exaggeration (e.g. 5)"),
    ("notes", "Figure notes (one line; use sidebar for long text)"),
)

COLLAR_COLUMNS = ("hole_id", "easting", "northing", "elevation", "total_depth")
LITHOLOGY_COLUMNS = ("hole_id", "from_depth", "to_depth", "lithology_code", "unit_order")
WATER_COLUMNS = ("hole_id", "depth", "elevation_masl", "series_id", "series_label")
ENVIRONMENTAL_COLUMNS = (
    "hole_id",
    "parameter",
    "value",
    "depth",
    "from_depth",
    "to_depth",
    "unit",
    "value_label",
)
SCREEN_COLUMNS = ("hole_id", "from_depth", "to_depth")
GRADIENT_COLUMNS = ("hole_id", "direction")

TABLE_SECTIONS: dict[str, tuple[str, ...]] = {
    "COLLARS": COLLAR_COLUMNS,
    "LITHOLOGY": LITHOLOGY_COLUMNS,
    "WATER": WATER_COLUMNS,
    "ENVIRONMENTAL": ENVIRONMENTAL_COLUMNS,
    "SCREENS": SCREEN_COLUMNS,
    "GRADIENTS": GRADIENT_COLUMNS,
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
_TITLE_FONT = Font(bold=True, name="Calibri", size=14, color="1F4E79")
_SECTION_FONT = Font(bold=True, name="Calibri", size=12, color="1F4E79")
_HINT_FONT = Font(italic=True, name="Calibri", size=10, color="666666")
_THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
_TAB_COLORS = {
    INSTRUCTIONS_SHEET: "4472C4",
    PROJECT_SHEET: "ED7D31",
    COLLARS_SHEET: "70AD47",
    LITHOLOGY_SHEET: "70AD47",
    WATER_SHEET: "5B9BD5",
    ENVIRONMENTAL_SHEET: "5B9BD5",
    SCREENS_SHEET: "A9D08E",
    GRADIENTS_SHEET: "A9D08E",
    EXAMPLE_SHEET: "7030A0",
    DATA_ENTRY_SHEET: "7F7F7F",
}


@dataclass(frozen=True)
class DataEntrySheets:
    project: dict[str, str]
    collars: pd.DataFrame
    lithology: pd.DataFrame
    water: pd.DataFrame
    environmental: pd.DataFrame
    screens: pd.DataFrame
    gradients: pd.DataFrame


def _sample_project() -> dict[str, str]:
    return {
        "client_name": "C-GROUP ENERGY INC.",
        "prepared_by": "ECOVENTURE",
        "project_number": "100/09-36-055-02 W4M",
        "section_title": "A - A' WITH CHLORIDE AVERAGES",
        "report_date": "06/24/26",
        "drawn_by": "SL/JG",
        "data_source": "ECOVENTURE 2026",
        "map_scale": "1:1000",
        "coordinate_reference": "EPSG:32611",
        "transect_start": "A / NORTHWEST",
        "transect_end": "A' / SOUTHEAST",
        "vertical_exaggeration": "5",
        "notes": "NOTE: masl DENOTES METRES ABOVE SEA LEVEL.",
    }


def _sample_collars() -> list[dict[str, object]]:
    return [
        {"hole_id": "MW-01", "easting": 0.0, "northing": 0.0, "elevation": 635.0, "total_depth": 12.0},
        {"hole_id": "MW-02", "easting": 45.0, "northing": 0.0, "elevation": 634.5, "total_depth": 12.0},
        {"hole_id": "MW-03", "easting": 90.0, "northing": 0.0, "elevation": 634.0, "total_depth": 11.5},
    ]


def _sample_lithology() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for hole_id, sand_to, clay_loam_to in (
        ("MW-01", 3.0, 8.0),
        ("MW-02", 3.0, 8.0),
        ("MW-03", 2.5, 7.5),
    ):
        total = 12.0 if hole_id != "MW-03" else 11.5
        rows.extend(
            [
                {
                    "hole_id": hole_id,
                    "from_depth": 0.0,
                    "to_depth": 0.5,
                    "lithology_code": "Topsoil",
                    "unit_order": "",
                },
                {
                    "hole_id": hole_id,
                    "from_depth": 0.5,
                    "to_depth": sand_to,
                    "lithology_code": "Sand",
                    "unit_order": "",
                },
                {
                    "hole_id": hole_id,
                    "from_depth": sand_to,
                    "to_depth": clay_loam_to,
                    "lithology_code": "Clay Loam",
                    "unit_order": "",
                },
                {
                    "hole_id": hole_id,
                    "from_depth": clay_loam_to,
                    "to_depth": total,
                    "lithology_code": "Clay",
                    "unit_order": "",
                },
            ]
        )
    return rows


def _sample_water() -> list[dict[str, object]]:
    return [
        {
            "hole_id": "MW-01",
            "depth": 3.2,
            "elevation_masl": "",
            "series_id": "2025-06",
            "series_label": "June 2025",
        },
        {
            "hole_id": "MW-02",
            "depth": "",
            "elevation_masl": 631.0,
            "series_id": "2025-06",
            "series_label": "June 2025",
        },
        {
            "hole_id": "MW-03",
            "depth": 3.8,
            "elevation_masl": "",
            "series_id": "2025-06",
            "series_label": "June 2025",
        },
    ]


def _sample_environmental() -> list[dict[str, object]]:
    return [
        {
            "hole_id": "MW-01",
            "parameter": "Chloride",
            "value": 5.0,
            "depth": 1.3,
            "from_depth": "",
            "to_depth": "",
            "unit": "mg/L",
            "value_label": "<5 mg/L",
        },
        {
            "hole_id": "MW-01",
            "parameter": "Chloride",
            "value": 12.2,
            "depth": 2.7,
            "from_depth": "",
            "to_depth": "",
            "unit": "mg/L",
            "value_label": "",
        },
        {
            "hole_id": "MW-02",
            "parameter": "Chloride",
            "value": 85.0,
            "depth": 3.5,
            "from_depth": "",
            "to_depth": "",
            "unit": "mg/L",
            "value_label": "",
        },
        {
            "hole_id": "MW-03",
            "parameter": "Chloride",
            "value": 120.0,
            "depth": 4.0,
            "from_depth": "",
            "to_depth": "",
            "unit": "mg/L",
            "value_label": "",
        },
        {
            "hole_id": "MW-02",
            "parameter": "Benzene",
            "value": 0.5,
            "depth": 2.0,
            "from_depth": "",
            "to_depth": "",
            "unit": "mg/L",
            "value_label": "",
        },
    ]


def _sample_screens() -> list[dict[str, object]]:
    return [
        {"hole_id": "MW-01", "from_depth": 8.0, "to_depth": 11.0},
        {"hole_id": "MW-02", "from_depth": 8.0, "to_depth": 11.0},
    ]


def _sample_gradients() -> list[dict[str, object]]:
    return [
        {"hole_id": "MW-01", "direction": "down"},
        {"hole_id": "MW-03", "direction": "up"},
    ]


def _column_hints() -> dict[str, dict[str, str]]:
    return {
        COLLARS_SHEET: {
            "hole_id": "Unique borehole ID (must match other sheets)",
            "easting": "Easting / X (metres)",
            "northing": "Northing / Y (metres)",
            "elevation": "Collar RL (masl)",
            "total_depth": "Total drilled depth (m below collar)",
        },
        LITHOLOGY_SHEET: {
            "hole_id": "Must match Collars.hole_id",
            "from_depth": "Top of interval (m below collar)",
            "to_depth": "Base of interval (m below collar)",
            "lithology_code": "e.g. Clay, Sand, Clay Loam, Topsoil",
            "unit_order": "Optional; leave blank to auto-assign from depth",
        },
        WATER_SHEET: {
            "hole_id": "Must match Collars.hole_id",
            "depth": "Water depth below collar (m) — OR use elevation_masl",
            "elevation_masl": "Water RL (masl) — do not fill both depth and this",
            "series_id": "Optional snapshot id (e.g. 2025-06)",
            "series_label": "Optional legend label (e.g. June 2025)",
        },
        ENVIRONMENTAL_SHEET: {
            "hole_id": "Must match Collars.hole_id",
            "parameter": "e.g. Chloride, Benzene",
            "value": "Numeric concentration (use detection limit for ND)",
            "depth": "Point sample depth (m) — OR use from/to interval",
            "from_depth": "Interval top (m) if not using depth",
            "to_depth": "Interval base (m) if not using depth",
            "unit": "e.g. mg/L",
            "value_label": "Optional figure text (e.g. <5 mg/L)",
        },
        SCREENS_SHEET: {
            "hole_id": "Must match Collars.hole_id",
            "from_depth": "Screen top (m below collar)",
            "to_depth": "Screen base (m below collar)",
        },
        GRADIENTS_SHEET: {
            "hole_id": "Must match Collars.hole_id",
            "direction": "up or down",
        },
    }


def _instructions_lines() -> list[str]:
    try:
        from constants import USGS_LITHOLOGY_COLORS

        lithology_list = ", ".join(sorted(USGS_LITHOLOGY_COLORS))
    except Exception:
        lithology_list = (
            "Clay, Sand, Clay Loam, Sandy Clay, Silty Clay, Sandstone, Silt, "
            "Sand and Gravel, Organics, Topsoil, Fill"
        )
    return [
        "CROSS SECTION STUDIO — INPUT WORKBOOK",
        "",
        "QUICK START",
        "1. Open the Project tab and replace the sample client / figure metadata.",
        "2. Enter boreholes on Collars (one row per hole).",
        "3. Enter lithology intervals on Lithology (from_depth / to_depth below collar).",
        "4. Optionally fill Water, Environmental, Screens, and Gradients.",
        "5. Upload this file in Cross Section Studio (Upload step).",
        "6. On Configure, pick the transect holes and which lab parameters to plot.",
        "",
        "TAB GUIDE",
        "• Instructions — this guide.",
        "• Project — client name, project number, section title, transect labels (seeds consulting layout).",
        "• Collars — required. Coordinates and collar elevation (RL).",
        "• Lithology — required. Stick-log intervals; hole_id must match Collars.",
        "• Water — optional. Groundwater as depth below collar OR elevation_masl (not both on one row).",
        "• Environmental — optional. Lab/field parameters at a point depth or depth interval.",
        "• Screens — optional. Screened intervals (consulting hatch bands).",
        "• Gradients — optional. Vertical gradient arrows (direction = up or down).",
        "• Example — filled MW-01 / MW-02 / MW-03 demo. Copy rows into the data tabs, or replace samples.",
        "• Data Entry — compatibility sheet (PROJECT metadata for auto-detect). Prefer the named tabs above.",
        "",
        "REQUIRED",
        "• Collars: hole_id, easting, northing, elevation, total_depth",
        "• Lithology: hole_id, from_depth, to_depth, lithology_code",
        "",
        "ENVIRONMENTAL RULES",
        "• Use depth for a point sample, OR from_depth + to_depth for an interval — not both.",
        "• value_label is optional display text on the figure (e.g. <5 mg/L for non-detects).",
        "",
        f"COMMON LITHOLOGY CODES: {lithology_list}",
        "",
        "TIPS",
        "• Keep hole_id spelling identical across every tab.",
        "• Depths are metres below collar unless the column is elevation_masl / elevation (RL).",
        "• Leave optional tabs blank (header only) if unused — do not delete the sheet.",
        "• Filter and freeze panes are enabled on data tabs for easier entry.",
        "• See docs/workbook-format.md for advanced sheets (Deviations, Correlations, Faults).",
    ]


def _blank_rows(count: int, columns: tuple[str, ...]) -> list[dict[str, object]]:
    return [{column: "" for column in columns} for _ in range(count)]


def _build_data_entry_project_only_rows() -> list[list[object]]:
    """Slim Data Entry sheet: PROJECT metadata for ingest + pointer to named tabs."""
    rows: list[list[object]] = [
        ["CROSS SECTION STUDIO — DATA ENTRY (compatibility)", "", "", "", "", "", "", ""],
        [
            "Prefer Collars / Lithology / Water / Environmental / Screens / Gradients tabs for data.",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "This sheet keeps PROJECT metadata for upload detection and consulting title seeding.",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        [""],
        ["PROJECT / CLIENT METADATA", "value", "", "", "", "", "", ""],
    ]
    project = _sample_project()
    for field, _label in PROJECT_FIELDS:
        rows.append([field, project.get(field, ""), "", "", "", "", "", ""])
    rows.append([""])
    rows.append(
        [
            "COLLARS (optional here — use Collars tab)",
            *([""] * (len(COLLAR_COLUMNS) - 1)),
        ]
    )
    rows.append(list(COLLAR_COLUMNS))
    for blank in _blank_rows(2, COLLAR_COLUMNS):
        rows.append([blank[column] for column in COLLAR_COLUMNS])
    rows.append([""])
    rows.append(
        [
            "LITHOLOGY (optional here — use Lithology tab)",
            *([""] * (len(LITHOLOGY_COLUMNS) - 1)),
        ]
    )
    rows.append(list(LITHOLOGY_COLUMNS))
    for blank in _blank_rows(2, LITHOLOGY_COLUMNS):
        rows.append([blank[column] for column in LITHOLOGY_COLUMNS])
    return rows


def _normalize_key(value: object) -> str:
    return str(value).strip().lower().replace(" ", "_")


def _table_frame(rows: list[list[object]], columns: tuple[str, ...]) -> pd.DataFrame:
    cleaned: list[dict[str, object]] = []
    for row in rows:
        if not row or all(str(cell).strip() == "" or str(cell).strip().lower() == "nan" for cell in row):
            continue
        payload = {columns[index]: row[index] if index < len(row) else "" for index in range(len(columns))}
        key_col = columns[0]
        if str(payload.get(key_col, "")).strip() == "":
            continue
        cleaned.append(payload)
    if not cleaned:
        return pd.DataFrame(columns=list(columns))
    return pd.DataFrame(cleaned)


def parse_data_entry_sheet(frame: pd.DataFrame) -> DataEntrySheets:
    """Parse the unified Data Entry layout into section DataFrames."""
    project: dict[str, str] = {}
    section_rows: dict[str, list[list[object]]] = {key: [] for key in TABLE_SECTIONS}
    mode = "scan"
    current_section: str | None = None
    project_fields = {field for field, _label in PROJECT_FIELDS}

    for row in frame.fillna("").itertuples(index=False, name=None):
        cells = list(row)
        first = str(cells[0]).strip() if cells else ""
        if not first:
            continue
        upper = first.upper()

        matched_section: str | None = None
        for key in TABLE_SECTIONS:
            if upper.startswith(key):
                matched_section = key
                break
        if matched_section:
            current_section = matched_section
            mode = "table_header"
            continue

        if upper.startswith("PROJECT"):
            mode = "project"
            current_section = None
            continue
        if first.lower() == "field":
            continue

        if mode == "project":
            key = _normalize_key(first)
            if key in project_fields and len(cells) >= 2:
                project[key] = str(cells[1]).strip()
            continue

        if mode == "table_header" and current_section:
            mode = "table"
            continue

        if mode == "table" and current_section:
            header_key = _normalize_key(TABLE_SECTIONS[current_section][0])
            if _normalize_key(first) == header_key:
                continue
            section_rows[current_section].append(cells)

    return DataEntrySheets(
        project=project,
        collars=_table_frame(section_rows["COLLARS"], COLLAR_COLUMNS),
        lithology=_table_frame(section_rows["LITHOLOGY"], LITHOLOGY_COLUMNS),
        water=_table_frame(section_rows["WATER"], WATER_COLUMNS),
        environmental=_table_frame(section_rows["ENVIRONMENTAL"], ENVIRONMENTAL_COLUMNS),
        screens=_table_frame(section_rows["SCREENS"], SCREEN_COLUMNS),
        gradients=_table_frame(section_rows["GRADIENTS"], GRADIENT_COLUMNS),
    )


def _style_header_row(worksheet, row: int, column_count: int) -> None:
    for col in range(1, column_count + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = _THIN


def _autosize_columns(worksheet, widths: dict[str, float] | None = None) -> None:
    if widths:
        for letter, width in widths.items():
            worksheet.column_dimensions[letter].width = width
        return
    for index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(index)
        max_len = 12
        for row in worksheet.iter_rows(min_col=index, max_col=index, values_only=True):
            value = row[0]
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)) + 2, 42))
        worksheet.column_dimensions[letter].width = max_len


def _write_dataframe_sheet(
    writer: pd.ExcelWriter,
    *,
    sheet_name: str,
    columns: tuple[str, ...],
    sample: list[dict[str, object]],
    blank_count: int,
    required: bool,
) -> None:
    frame = pd.DataFrame(sample, columns=list(columns))
    frame.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.book[sheet_name]
    worksheet.sheet_properties.tabColor = _TAB_COLORS.get(sheet_name)
    _style_header_row(worksheet, 1, len(columns))
    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 30
    hints = _column_hints().get(sheet_name, {})
    for index, column in enumerate(columns, start=1):
        hint = hints.get(column)
        if hint:
            worksheet.cell(row=1, column=index).comment = Comment(
                hint, "Cross Section Studio", width=240, height=60
            )
    # Reserve empty rows for typing without writing placeholder values the parser would read.
    last_data_row = 1 + len(sample)
    entry_end = last_data_row + max(blank_count, 20)
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{entry_end}"
    status = "REQUIRED — replace sample rows with your boreholes" if required else (
        "OPTIONAL — leave blank if unused (do not delete this sheet)"
    )
    # Place the hint to the right of the table so it is never read as hole_id.
    hint_col = len(columns) + 2
    worksheet.cell(row=1, column=hint_col, value=status).font = _HINT_FONT
    _autosize_columns(worksheet)
    if sheet_name == GRADIENTS_SHEET:
        validation = DataValidation(
            type="list",
            formula1='"up,down"',
            allow_blank=True,
            showDropDown=False,
        )
        validation.error = "Use up or down"
        validation.errorTitle = "Gradient direction"
        worksheet.add_data_validation(validation)
        validation.add(f"B2:B{entry_end}")


def _write_project_sheet(writer: pd.ExcelWriter) -> None:
    project = _sample_project()
    rows = [
        {"field": field, "description": label, "value": project.get(field, "")}
        for field, label in PROJECT_FIELDS
    ]
    frame = pd.DataFrame(rows)
    frame.to_excel(writer, sheet_name=PROJECT_SHEET, index=False)
    worksheet = writer.book[PROJECT_SHEET]
    worksheet.sheet_properties.tabColor = _TAB_COLORS[PROJECT_SHEET]
    _style_header_row(worksheet, 1, 3)
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 48
    worksheet.column_dimensions["C"].width = 42
    worksheet.cell(row=len(rows) + 3, column=1, value="Edit the value column only.").font = _HINT_FONT
    worksheet.cell(
        row=len(rows) + 4,
        column=1,
        value="These fields seed the consulting report title block on upload.",
    ).font = _HINT_FONT


def _write_instructions_sheet(writer: pd.ExcelWriter) -> None:
    instructions = pd.DataFrame({"Instructions": _instructions_lines()})
    instructions.to_excel(writer, sheet_name=INSTRUCTIONS_SHEET, index=False)
    worksheet = writer.book[INSTRUCTIONS_SHEET]
    worksheet.sheet_properties.tabColor = _TAB_COLORS[INSTRUCTIONS_SHEET]
    worksheet.column_dimensions["A"].width = 110
    worksheet.cell(row=1, column=1).font = _HEADER_FONT
    worksheet.cell(row=1, column=1).fill = _HEADER_FILL
    for row_index, line in enumerate(_instructions_lines(), start=2):
        cell = worksheet.cell(row=row_index, column=1)
        if line in {
            "CROSS SECTION STUDIO — INPUT WORKBOOK",
            "QUICK START",
            "TAB GUIDE",
            "REQUIRED",
            "ENVIRONMENTAL RULES",
            "TIPS",
        } or line.startswith("COMMON LITHOLOGY"):
            cell.font = _SECTION_FONT if line != "CROSS SECTION STUDIO — INPUT WORKBOOK" else _TITLE_FONT


def _write_example_sheet(writer: pd.ExcelWriter) -> None:
    """Reference layout showing the sample project as labeled tables (not parsed)."""
    rows: list[list[object]] = [
        ["EXAMPLE PROJECT — MW-01 / MW-02 / MW-03 (reference only)", "", "", "", "", "", "", ""],
        [
            "Copy rows into Collars / Lithology / Water / Environmental / Screens / Gradients, or replace the sample rows already on those tabs.",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        [""],
        ["PROJECT / CLIENT METADATA", "value", "", "", "", "", "", ""],
    ]
    project = _sample_project()
    for field, _label in PROJECT_FIELDS:
        rows.append([field, project.get(field, ""), "", "", "", "", "", ""])
    rows.append([""])

    def append_table(section: str, columns: tuple[str, ...], sample: list[dict[str, object]]) -> None:
        rows.append([section, *([""] * max(len(columns) - 1, 0))])
        rows.append(list(columns))
        for item in sample:
            rows.append([item.get(column, "") for column in columns])
        rows.append([""])

    append_table("COLLARS (required)", COLLAR_COLUMNS, _sample_collars())
    append_table("LITHOLOGY (required)", LITHOLOGY_COLUMNS, _sample_lithology())
    append_table("WATER (optional)", WATER_COLUMNS, _sample_water())
    append_table("ENVIRONMENTAL / LAB (optional)", ENVIRONMENTAL_COLUMNS, _sample_environmental())
    append_table("SCREENS (optional)", SCREEN_COLUMNS, _sample_screens())
    append_table("GRADIENTS (optional)", GRADIENT_COLUMNS, _sample_gradients())

    frame = pd.DataFrame(rows)
    frame.to_excel(writer, sheet_name=EXAMPLE_SHEET, index=False, header=False)
    worksheet = writer.book[EXAMPLE_SHEET]
    worksheet.sheet_properties.tabColor = _TAB_COLORS[EXAMPLE_SHEET]
    worksheet.cell(row=1, column=1).font = _TITLE_FONT
    worksheet.cell(row=2, column=1).font = _HINT_FONT
    worksheet.column_dimensions["A"].width = 28
    for column in "BCDEFGH":
        worksheet.column_dimensions[column].width = 14


def _write_data_entry_compat_sheet(writer: pd.ExcelWriter) -> None:
    data_entry = pd.DataFrame(_build_data_entry_project_only_rows())
    data_entry.to_excel(writer, sheet_name=DATA_ENTRY_SHEET, index=False, header=False)
    worksheet = writer.book[DATA_ENTRY_SHEET]
    worksheet.sheet_properties.tabColor = _TAB_COLORS[DATA_ENTRY_SHEET]
    worksheet.column_dimensions["A"].width = 56
    for column in "BCDEFGH":
        worksheet.column_dimensions[column].width = 14
    worksheet.cell(row=1, column=1).font = _SECTION_FONT
    worksheet.cell(row=2, column=1).font = _HINT_FONT
    worksheet.cell(row=3, column=1).font = _HINT_FONT


def build_input_template(path: Path | None = None) -> Path:
    """Write the multi-tab input template workbook."""
    from paths import build_input_template_path

    output = path or build_input_template_path()
    output.parent.mkdir(parents=True, exist_ok=True)

    try:
        return _write_input_template(output)
    except PermissionError:
        if path is not None:
            raise
        from paths import app_root

        fallback = app_root() / "data" / "Cross_Section_Input_Template_v3.xlsx"
        return _write_input_template(fallback)


def _populate_input_template(writer: pd.ExcelWriter) -> None:
    """Write all input-template sheets into an open ExcelWriter."""
    _write_instructions_sheet(writer)
    _write_project_sheet(writer)
    _write_dataframe_sheet(
        writer,
        sheet_name=COLLARS_SHEET,
        columns=COLLAR_COLUMNS,
        sample=_sample_collars(),
        blank_count=5,
        required=True,
    )
    _write_dataframe_sheet(
        writer,
        sheet_name=LITHOLOGY_SHEET,
        columns=LITHOLOGY_COLUMNS,
        sample=_sample_lithology(),
        blank_count=8,
        required=True,
    )
    _write_dataframe_sheet(
        writer,
        sheet_name=WATER_SHEET,
        columns=WATER_COLUMNS,
        sample=_sample_water(),
        blank_count=5,
        required=False,
    )
    _write_dataframe_sheet(
        writer,
        sheet_name=ENVIRONMENTAL_SHEET,
        columns=ENVIRONMENTAL_COLUMNS,
        sample=_sample_environmental(),
        blank_count=8,
        required=False,
    )
    _write_dataframe_sheet(
        writer,
        sheet_name=SCREENS_SHEET,
        columns=SCREEN_COLUMNS,
        sample=_sample_screens(),
        blank_count=5,
        required=False,
    )
    _write_dataframe_sheet(
        writer,
        sheet_name=GRADIENTS_SHEET,
        columns=GRADIENT_COLUMNS,
        sample=_sample_gradients(),
        blank_count=5,
        required=False,
    )
    _write_example_sheet(writer)
    _write_data_entry_compat_sheet(writer)


def _write_input_template(output: Path) -> Path:
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _populate_input_template(writer)
    return output


def build_input_template_bytes() -> bytes:
    """Return the multi-tab input template as .xlsx bytes (Cloud / in-app download)."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _populate_input_template(writer)
    return buffer.getvalue()


def load_data_entry_sheets(source: str | Path | BinaryIO | BytesIO) -> DataEntrySheets:
    frame = pd.read_excel(source, sheet_name=DATA_ENTRY_SHEET, header=None)
    return parse_data_entry_sheet(frame)


def _load_project_sheet_metadata(source: str | Path | BinaryIO | BytesIO) -> dict[str, str]:
    frame = pd.read_excel(source, sheet_name=PROJECT_SHEET)
    normalized = {str(col).strip().lower(): col for col in frame.columns}
    field_col = normalized.get("field")
    value_col = normalized.get("value")
    if field_col is None or value_col is None:
        return {}
    project_fields = {field for field, _label in PROJECT_FIELDS}
    result: dict[str, str] = {}
    for _, row in frame.iterrows():
        key = _normalize_key(row[field_col])
        if key in project_fields:
            value = str(row[value_col]).strip()
            if value and value.lower() != "nan":
                result[key] = value
    return result


def load_project_metadata(source: str | Path | BinaryIO | BytesIO) -> dict[str, str]:
    """Load PROJECT fields — Project tab overrides Data Entry field-by-field."""
    from_data_entry: dict[str, str] = {}
    try:
        if hasattr(source, "seek"):
            source.seek(0)
        from_data_entry = dict(load_data_entry_sheets(source).project)
    except Exception:
        from_data_entry = {}
    from_project: dict[str, str] = {}
    try:
        if hasattr(source, "seek"):
            source.seek(0)
        from_project = _load_project_sheet_metadata(source)
    except Exception:
        from_project = {}
    merged = dict(from_data_entry)
    merged.update(from_project)
    return merged
